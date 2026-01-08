# -*- coding: utf-8 -*-
"""
分析 Excel 文件结构
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def build_merged_cells_map(ws):
    """构建合并单元格映射"""
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_map[(row, col)] = (min_row, min_col)
    
    return merged_map


def get_cell_value(ws, row, col, merged_map):
    """获取单元格值（处理合并）"""
    cell_coord = (row, col)
    if cell_coord in merged_map:
        master_coord = merged_map[cell_coord]
        return ws.cell(master_coord[0], master_coord[1]).value
    else:
        return ws.cell(row, col).value


def analyze_excel_headers(file_path):
    """分析 Excel 表头结构"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    merged_map = build_merged_cells_map(ws)
    max_col = ws.max_column
    
    print("=" * 120)
    print("📊 Excel 表头结构分析")
    print("=" * 120)
    print()
    
    # 打印第1行和第2行表头
    print("列号 | 列字母 | 第1行表头 | 第2行表头")
    print("-" * 120)
    
    for col in range(1, min(max_col + 1, 160)):  # 限制最多160列
        col_letter = get_column_letter(col)
        val1 = get_cell_value(ws, 1, col, merged_map)
        val2 = get_cell_value(ws, 2, col, merged_map)
        
        val1_str = str(val1)[:20] if val1 else ""
        val2_str = str(val2)[:20] if val2 else ""
        
        print(f"{col:4d} | {col_letter:6s} | {val1_str:30s} | {val2_str:30s}")
    
    print()
    print("=" * 120)
    print(f"总列数: {max_col}")
    print(f"总行数: {ws.max_row}")
    print("=" * 120)
    
    # 显示前几行数据示例
    print()
    print("📋 前5行数据示例（前10列）：")
    print("=" * 120)
    for row in range(1, min(8, ws.max_row + 1)):
        print(f"第 {row} 行:")
        for col in range(1, min(11, max_col + 1)):
            val = get_cell_value(ws, row, col, merged_map)
            print(f"  {get_column_letter(col):3s}: {str(val)[:40]}")
        print()


if __name__ == "__main__":
    file_name = "副本（现RPA小工具流程）简历导入多行表.xlsx"
    analyze_excel_headers(file_name)

