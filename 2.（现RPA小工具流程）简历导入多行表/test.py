from openpyxl import load_workbook
import json

wb = load_workbook('简历导入多行表_和并单元格测试.xlsx')
ws = wb.active

merged_info = []
for merged_cell in ws.merged_cells.ranges:
    merged_info.append({
        "range": str(merged_cell),
        "start_row": merged_cell.min_row,
        "end_row": merged_cell.max_row,
        "rows_merged": merged_cell.max_row - merged_cell.min_row + 1,
        "is_merged": True
    })

json_output = json.dumps(merged_info, ensure_ascii=False, indent=2)
print(json_output)